import os
import logging
import threading
import traceback

from flask import Flask, request

import pdfplumber
from openpyxl import Workbook

from telegram import Bot, Update
from telegram.ext import Dispatcher, MessageHandler, CommandHandler, Filters

# -------------------------------------------------------------------
# Logging setup
# -------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
)
logger = logging.getLogger(__name__)

# -------------------------------------------------------------------
# Telegram / Flask setup
# -------------------------------------------------------------------
TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise RuntimeError("BOT_TOKEN env var is not set")

bot = Bot(TOKEN)
app = Flask(__name__)

dispatcher = Dispatcher(bot, None, workers=0, use_context=True)

# Active jobs: chat_id -> {"cancel_event": Event, "thread": Thread}
active_jobs = {}
jobs_lock = threading.Lock()

# Limits – tune these for Render free tier
MAX_PAGES = 3000                  # hard safety limit on pages to process
MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024  # 20 MB – reject bigger PDFs
PROGRESS_STEPS = 10               # how many times to update progress


# -------------------------------------------------------------------
# Background worker: PDF -> Excel with low memory use
# -------------------------------------------------------------------
def process_pdf_async(
    chat_id: int,
    status_message_id: int,
    pdf_path: str,
    excel_path: str,
    cancel_event: threading.Event,
) -> None:
    """
    Runs in a background thread.
    Reads PDF page-by-page and writes directly to an Excel workbook.
    Avoids building huge lists in RAM.
    """
    logger.info("Starting background job for chat_id=%s, pdf=%s", chat_id, pdf_path)
    try:
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=status_message_id,
            text="Processing your PDF… ⏳",
        )

        # Create workbook in memory (more efficient than pandas DataFrame for big data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        total_pages_to_process = 0

        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 0:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=status_message_id,
                    text="⚠️ This PDF seems to be empty.",
                )
                return

            if total_pages > MAX_PAGES:
                logger.warning(
                    "PDF too large: %s pages for chat_id=%s. Limiting to %s pages.",
                    total_pages,
                    chat_id,
                    MAX_PAGES,
                )

            total_pages_to_process = min(total_pages, MAX_PAGES)

            step = max(1, total_pages_to_process // PROGRESS_STEPS)

            for i in range(total_pages_to_process):
                if cancel_event.is_set():
                    logger.info("Job cancelled by user, chat_id=%s", chat_id)
                    bot.edit_message_text(
                        chat_id=chat_id,
                        message_id=status_message_id,
                        text=f"❌ Conversion cancelled at page {i}/{total_pages_to_process}.",
                    )
                    return

                page_num = i + 1
                page = pdf.pages[i]

                try:
                    table = page.extract_table()
                except Exception as e:
                    logger.exception(
                        "Error extracting table on page %s for chat_id=%s: %s",
                        page_num,
                        chat_id,
                        e,
                    )
                    table = None

                if table:
                    for row in table:
                        # row is a list of cell values
                        ws.append(row)

                # Update progress
                if page_num % step == 0 or page_num == total_pages_to_process:
                    percent = int(page_num * 100 / total_pages_to_process)
                    try:
                        bot.edit_message_text(
                            chat_id=chat_id,
                            message_id=status_message_id,
                            text=(
                                f"Processing… {percent}% done ⏳ "
                                f"({page_num}/{total_pages_to_process} pages)"
                            ),
                        )
                    except Exception as e:
                        logger.warning("Failed to edit progress message: %s", e)

        # If no rows wrote to sheet (only header exists or nothing at all)
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=status_message_id,
                text="⚠️ I couldn't detect any tables in this PDF.",
            )
            return

        # Save workbook to disk
        wb.save(excel_path)

        bot.edit_message_text(
            chat_id=chat_id,
            message_id=status_message_id,
            text="✅ Conversion complete! Sending your Excel file…",
        )

        with open(excel_path, "rb") as f:
            bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=os.path.basename(excel_path),
                caption="Here is your converted Excel file 😊",
            )

        logger.info("Job completed successfully for chat_id=%s", chat_id)

    except Exception as e:
        logger.error("Unhandled error in background job: %s", e)
        logger.error(traceback.format_exc())
        try:
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=status_message_id,
                text="❌ An error occurred while converting your PDF. Please try again later.",
            )
        except Exception as inner_e:
            logger.warning("Failed to edit message after error: %s", inner_e)
    finally:
        # Cleanup
        with jobs_lock:
            active_jobs.pop(chat_id, None)
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            if os.path.exists(excel_path):
                os.remove(excel_path)
        except Exception as e:
            logger.warning("Failed to delete temp files: %s", e)


# -------------------------------------------------------------------
# Handlers
# -------------------------------------------------------------------
def start(update, context):
    chat_id = update.effective_chat.id
    logger.info("/start from chat_id=%s", chat_id)
    update.message.reply_text(
        "Hi! 👋\n\n"
        "Send me a *PDF with tables* and I’ll convert it to Excel 📊\n"
        "You can send /stop to cancel a running conversion.\n\n"
        "_Note: very large PDFs may be partially processed due to server limits._",
        parse_mode="Markdown",
    )


def stop(update, context):
    chat_id = update.effective_chat.id
    logger.info("/stop from chat_id=%s", chat_id)

    with jobs_lock:
        job = active_jobs.get(chat_id)

    if not job:
        update.message.reply_text("There is no active conversion to stop 🙂")
        return

    job["cancel_event"].set()
    update.message.reply_text("🛑 Stop requested. I’ll cancel the current conversion.")


def handle_pdf(update, context):
    message = update.message
    chat_id = message.chat.id
    document = message.document

    logger.info(
        "Received document from chat_id=%s: %s (%s bytes)",
        chat_id,
        document.file_name,
        document.file_size,
    )

    if not document.file_name.lower().endswith(".pdf"):
        message.reply_text("Please upload a PDF file 😄")
        return

    # File size limit to avoid OOM on Render
    if document.file_size and document.file_size > MAX_FILE_SIZE_BYTES:
        mb = MAX_FILE_SIZE_BYTES // (1024 * 1024)
        message.reply_text(
            f"⚠️ This PDF is too large for this bot (limit ~{mb} MB).\n"
            "Please split the file or process it in parts."
        )
        logger.warning(
            "Rejected file from chat_id=%s due to size: %s bytes",
            chat_id,
            document.file_size,
        )
        return

    with jobs_lock:
        existing = active_jobs.get(chat_id)
        if existing:
            message.reply_text(
                "⚠️ You already have a conversion running.\n"
                "Send /stop to cancel it before starting a new one."
            )
            return

    # Download PDF to /tmp
    pdf_path = f"/tmp/{document.file_name}"
    excel_path = pdf_path.replace(".pdf", ".xlsx")

    try:
        file = document.get_file()
        file.download(pdf_path)
        logger.info("Downloaded PDF to %s for chat_id=%s", pdf_path, chat_id)
    except Exception as e:
        logger.error("Error downloading file: %s", e)
        message.reply_text("❌ Failed to download the PDF from Telegram.")
        return

    status_message = message.reply_text("Starting PDF processing… ⏳")
    status_message_id = status_message.message_id

    cancel_event = threading.Event()
    thread = threading.Thread(
        target=process_pdf_async,
        args=(chat_id, status_message_id, pdf_path, excel_path, cancel_event),
        daemon=True,
    )

    with jobs_lock:
        active_jobs[chat_id] = {
            "cancel_event": cancel_event,
            "thread": thread,
        }

    thread.start()
    logger.info("Background thread started for chat_id=%s", chat_id)


# -------------------------------------------------------------------
# Register handlers
# -------------------------------------------------------------------
dispatcher.add_handler(CommandHandler("start", start))
dispatcher.add_handler(CommandHandler("stop", stop))
dispatcher.add_handler(MessageHandler(Filters.document.pdf, handle_pdf))


# -------------------------------------------------------------------
# Flask routes
# -------------------------------------------------------------------
@app.route("/")
def home():
    return "🚀 Telegram PDF→Excel bot is running."


@app.route(f"/webhook/{TOKEN}", methods=["POST"])
def webhook():
    try:
        json_update = request.get_json(force=True, silent=True)
        if not json_update:
            logger.warning("Received empty update")
            return "no update", 200

        update = Update.de_json(json_update, bot)
        logger.info("Received update: update_id=%s", update.update_id)

        dispatcher.process_update(update)
    except Exception as e:
        logger.error("Error in webhook handler: %s", e)
        logger.error(traceback.format_exc())
    return "ok", 200


if __name__ == "__main__":
    port = int(os.getenv("PORT", "10000"))
    app.run(host="0.0.0.0", port=port, debug=True)
